from openpyxl import Workbook
from openpyxl.styles import Font
from calculate import *

import config

wb = Workbook()
ws = wb.active
font_new = Font(name='Times New Roman', size=10)
font_song = Font(name='宋体', size=10)

# q1填充行名
q1_dis_row_names = ["龙头x (m)", "龙头y (m)"]

for i in range(1, 222):
    q1_dis_row_names.append(f"第{i}节龙身x (m)")
    q1_dis_row_names.append(f"第{i}节龙身y (m)")

q1_dis_row_names.append("龙尾x (m)")
q1_dis_row_names.append("龙尾y (m)")
q1_dis_row_names.append("龙尾（后）x (m)")
q1_dis_row_names.append("龙尾（后）y (m)")

q1_v_row_names = ["龙头 (m/s)"]

for i in range(1, 222):
    q1_v_row_names.append(f"第{i}节龙身  (m/s)")

q1_v_row_names.append("龙尾  (m/s)")
q1_v_row_names.append("龙尾（后） (m/s)")

# q1填充列名
q1_column_names = [f"{i} s" for i in range(301)]  # 从 0 s 到 21 s

# q2填充行名
q2_row_names = ["龙头"]

for i in range(1, 222):
    q2_row_names.append(f"第{i}节龙身")

q2_row_names.append("龙尾")
q2_row_names.append("龙尾（后）")

# q2填充列名
q2_column_names = ["横坐标x (m)", "纵坐标y (m)", "速度 (m/s)"]

# q4_dis填充行名
q4_dis_row_names = q1_dis_row_names

# q4_dis填充列名
q4_dis_column_names = [f"{i} s" for i in range(-100, 101)]

# q4_v填充行名
q4_v_row_names = q1_v_row_names

# q4_v填充列名
q4_v_column_names = [f"{i} s" for i in range(-100, 101)]

def dis_xlsx_init():
    for row_num, row_name in enumerate(q1_dis_row_names, start=2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.font = font_new

    for col_num, col_name in enumerate(q1_column_names, start=2):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = font_new


def dis_fill_xlsx():
    for time in range(301):
        theta = t_to_theta(time)
        pos = get_actual_position(theta, config.space)
        x = pos[0][0] / 100.0
        y = pos[1][0] / 100.0
        cell = ws.cell(row=2, column=time + 2, value=f"{x:.6f}")
        cell.font = font_new
        cell = ws.cell(row=3, column=time + 2, value=f"{y:.6f}")
        cell.font = font_new

        # 更新接下来的点的位置
        current_theta = theta
        for i in range(1, 224):
            delta_theta = find_actual_delta_theta(current_theta, config.actual_fixed_distances[0 if i == 1 else 1], config.space)
            current_theta += delta_theta
            pos = get_actual_position(current_theta, config.space)
            x = pos[0][0] / 100.0
            y = pos[1][0] / 100.0
            print(f"{time} {i}")
            cell = ws.cell(row=2 * (i + 1), column=time + 2, value=f"{x:.6f}")
            cell.font = font_new
            cell = ws.cell(row=2 * (i + 1) + 1, column=time + 2, value=f"{y:.6f}")
            cell.font = font_new

    wb.save("result1_dis.xlsx")


def v_xlsx_init():
    for row_num, row_name in enumerate(q1_v_row_names, start=2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.font = font_new

    for col_num, col_name in enumerate(q1_column_names, start=2):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = font_new

def v_fill_xlsx():
    for time in range(301):
        dt = 1e-5
        theta = t_to_theta(time)
        pos = get_actual_position(theta, config.space)
        _theta = t_to_theta(time - dt)
        _pos = get_actual_position(_theta, config.space)
        ds = list_cartesian_distance(pos, _pos) / 100.0
        print(ds)
        v = ds / dt
        cell = ws.cell(row=2, column=time + 2, value=f"{v[0]:.6f}")
        cell.font = font_new

        # 更新接下来的点的位置
        current_theta = theta
        _current_theta = _theta
        for i in range(1, 224):
            delta_theta = find_actual_delta_theta(current_theta, config.actual_fixed_distances[0 if i == 1 else 1], config.space)
            current_theta += delta_theta
            pos = get_actual_position(current_theta, config.space)
            _delta_theta = find_actual_delta_theta(_current_theta, config.actual_fixed_distances[0 if i == 1 else 1], config.space)
            _current_theta += _delta_theta
            _pos = get_actual_position(_current_theta, config.space)
            ds = list_cartesian_distance(pos, _pos) / 100.0
            v = ds / dt
            print(time)
            cell = ws.cell(row=i + 2, column=time + 2, value=f"{v[0]:.6f}")
            cell.font = font_new

    wb.save("result1_v.xlsx")

def collision_xlsx_init():
    for row_num, row_name in enumerate(q2_row_names, start=2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.font = font_song if row_num == 2 else font_new

    for col_num, col_name in enumerate(q2_column_names, start=2):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = font_new

def collision_fill_xlsx():
    time = 412.48
    dt = 1e-5
    theta = t_to_theta(time)
    pos = get_actual_position(theta, config.space)
    x = pos[0][0] / 100.0
    y = pos[1][0] / 100.0
    cell = ws.cell(row=2, column=2, value=f"{x:.6f}")
    cell.font = font_new
    cell = ws.cell(row=2, column=3, value=f"{y:.6f}")
    cell.font = font_new
    _theta = t_to_theta(time - dt)
    _pos = get_actual_position(_theta, config.space)
    ds = list_cartesian_distance(pos, _pos) / 100.0
    v = ds / dt
    cell = ws.cell(row=2, column=4, value=f"{v[0]:.6f}")
    cell.font = font_new

    # 更新接下来的点的位置
    current_theta = theta
    _current_theta = _theta
    for i in range(1, 224):
        delta_theta = find_actual_delta_theta(current_theta, config.actual_fixed_distances[0 if i == 1 else 1],
                                              config.space)
        current_theta += delta_theta
        pos = get_actual_position(current_theta, config.space)
        x = pos[0][0] / 100.0
        y = pos[1][0] / 100.0
        cell = ws.cell(row=i + 2, column=2, value=f"{x:.6f}")
        cell.font = font_new
        cell = ws.cell(row=i + 2, column=3, value=f"{y:.6f}")
        cell.font = font_new
        _delta_theta = find_actual_delta_theta(_current_theta, config.actual_fixed_distances[0 if i == 1 else 1],
                                               config.space)
        _current_theta += _delta_theta
        _pos = get_actual_position(_current_theta, config.space)
        ds = list_cartesian_distance(pos, _pos) / 100.0
        v = ds / dt
        print(time)
        cell = ws.cell(row=i + 2, column=4, value=f"{v[0]:.6f}")
        cell.font = font_new

    wb.save("result2.xlsx")

def q4_dis_xlsx_init():
    for row_num, row_name in enumerate(q4_dis_row_names, start=2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.font = font_song if row_num == 2 else font_new

    for col_num, col_name in enumerate(q4_dis_column_names, start=2):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = font_new

def q4_dis_fill():
    for time in range(-100, 101):
        # 更新第一个点的位置
        x, y = t_to_xy_q4(time)
        cell = ws.cell(row=2, column=time + 102, value=f"{x:.6f}")
        cell.font = font_new
        cell = ws.cell(row=3, column=time + 102, value=f"{y:.6f}")
        cell.font = font_new

        ergodic_time = time
        for i in range(1, 224):
            delta_time = find_delta_time(ergodic_time, config.actual_fixed_distances[0 if i == 1 else 1])
            ergodic_time -= delta_time
            _x, _y = t_to_xy_q4(ergodic_time)
            cell = ws.cell(row=2 * (i + 1), column=time + 102, value=f"{_x:.6f}")
            cell.font = font_new
            cell = ws.cell(row=2 * (i + 1) + 1, column=time + 102, value=f"{_y:.6f}")
            cell.font = font_new

        print(time)

    wb.save("result4_dis.xlsx")

def q4_v_xlsx_init():
    for row_num, row_name in enumerate(q4_v_row_names, start=2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.font = font_song if row_num == 2 else font_new

    for col_num, col_name in enumerate(q4_v_column_names, start=2):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = font_new

def q4_v_fill():
    for time in range(-100, 101):
        dt = 1e-5
        x, y = t_to_xy_q4(time)
        _x, _y = t_to_xy_q4(time - dt)
        ds = list_cartesian_distance([x, y], [_x, _y])
        v = ds / dt
        cell = ws.cell(row=2, column=time + 102, value=f"{v:.6f}")
        cell.font = font_new

        # 更新接下来的点的位置
        ergodic_time = time
        _ergodic_time = time - dt
        for i in range(1, 224):
            delta_time = find_delta_time(ergodic_time, config.actual_fixed_distances[0 if i == 1 else 1])
            ergodic_time -= delta_time
            x, y = t_to_xy_q4(ergodic_time)
            _delta_time = find_delta_time(_ergodic_time, config.actual_fixed_distances[0 if i == 1 else 1])
            _ergodic_time -= _delta_time
            _x, _y = t_to_xy_q4(_ergodic_time)
            ds = list_cartesian_distance([x, y], [_x, _y])
            # print(i, ds)
            v = ds / dt
            cell = ws.cell(row=i + 2, column=time + 102, value=f"{v:.6f}")
            cell.font = font_new

        print(time)

    wb.save("result4_v.xlsx")


# dis_xlsx_init()
# dis_fill_xlsx()

# v_xlsx_init()
# v_fill_xlsx()

# collision_xlsx_init()
# collision_fill_xlsx()

# q4_dis_xlsx_init()
# q4_dis_fill()

# q4_v_xlsx_init()
# q4_v_fill()
