import random

import numpy as np
from astropy.modeling.functional_models import Box1D
from joblib.testing import param
from mpmath.math2 import sqrt2
from numpy.matrixlib.defmatrix import matrix
from pandas.core.indexes.multi import sparsify_labels
from statsmodels.tsa.statespace.simulation_smoother import check_random_state
from openpyxl import load_workbook
from sympy.abc import epsilon

import config
from calculate import t_to_dis
from calculate import t_to_theta

# 2 dot -> 4 edge
lambda_of_head = 27.5 / 286
lambda_of_not_head = 27.5 / 165
lambda_move_head = 15 / 286
lambda_move_not_head = 15 / 165
def cal_long_side(front_dot, back_dot, head):
    back_front_vector = front_dot - back_dot
    if head:
        front_end = np.array([(lambda_of_head + 1) * front_dot[0] - lambda_of_head * back_dot[0],
                              (lambda_of_head + 1) * front_dot[1] - lambda_of_head * back_dot[1]])
        back_end  = np.array([(lambda_of_head + 1) * back_dot[0] - lambda_of_head * front_dot[0],
                              (lambda_of_head + 1) * back_dot[1] - lambda_of_head * front_dot[1]])
        move_vector = np.array([back_front_vector[1] * lambda_move_head, - back_front_vector[0] * lambda_move_head])
    else :
        front_end = np.array([(lambda_of_not_head + 1) * front_dot[0] - lambda_of_not_head * back_dot[0],
                              (lambda_of_not_head + 1) * front_dot[1] - lambda_of_not_head * back_dot[1]])
        back_end  = np.array([(lambda_of_not_head + 1) * back_dot[0] - lambda_of_not_head * front_dot[0],
                              (lambda_of_not_head + 1) * back_dot[1] - lambda_of_not_head * front_dot[1]])
        move_vector = np.array([back_front_vector[1] * lambda_move_not_head, - back_front_vector[0] * lambda_move_not_head])
    # print(front_end + move_vector, back_end + move_vector, front_end - move_vector, back_end - move_vector)
    return front_end + move_vector, back_end + move_vector, front_end - move_vector, back_end - move_vector

# cal_long_side(np.array([0,0]), np.array([165 * sqrt2 / 2, 165 * sqrt2 / 2]), False)


# return true if 2 edges cross
def cross_test(edge1, edge2):
    A0 = edge1[0]
    A1 = edge1[1]
    B0 = edge2[0]
    B1 = edge2[1]

    A0B0 = B0 - A0
    A0B1 = B1 - A0
    A0A1 = A1 - A0

    vector1 = np.cross(A0A1, A0B0)
    vector2 = np.cross(A0A1, A0B1)

    B0A0 = -A0B0
    B0A1 = A1 - B0
    B0B1 = B1 - B0

    vector3 = np.cross(B0B1, B0A0)
    vector4 = np.cross(B0B1, B0A1)

    #print(vector1, vector2)
    #print(vector3, vector4)
    if vector1 * vector2 <= 0 and vector3 * vector4 <= 0:
        print(f'crossed edges:\n1:{edge1},\n2:{edge2}')
        #print(vector1, vector2, vector3, vector4)
    return vector1 * vector2 <= 0 and vector3 * vector4 <= 0

#print(cross_test(np.array([[0,0],[1,0]]),np.array([[0,-1],[0.000,0.3]])))

def all_cross_check(dot_matrix):
    edge_list = []
    for i in range(dot_matrix.shape[0] - 1):
        A0, A1, B0, B1 = cal_long_side(dot_matrix[i], dot_matrix[i + 1], i == 0)
        edge_list.append(np.array([A0, A1])) # even：inside
        edge_list.append(np.array([B0, B1])) # odd: outside

    cross = False
    for i in range(2, len(edge_list), 2):
        for j in range(i - 3, 0, -2):
            #print(i, j)
            if cross_test(edge_list[i], edge_list[j]):
                print(f'crossed benches:{i // 2}, {j // 2}')
                cross = True
                break
        if cross:
            break

    return cross

#print(all_cross_check(np.array([[0,0],[386,0],[386 + 165, 0]])))
#
# dot_matrix = np.empty((300, 224, 2))
# def read_from_dis_excel(time):
#      # 从表格读取，后面改成从程序读取
#      file_path = '../result1_dis.xlsx'
#      wb = load_workbook(file_path)
#
#
#      ws = wb.active
#      rows_to_read = range(2, 450)
#      cols_to_read = range(time + 2, time + 3)
#
#
#      for row_num in rows_to_read:
#          for col_num in  cols_to_read:
#              cell_value = ws.cell(row=row_num, column=col_num).value
#              #print(f'Cell ({row_num}, {col_num}): {cell_value}')
#              dot_matrix[time][row_num // 2 - 1][row_num % 2] = cell_value
#
#      #print(dot_matrix)

# def check_excel():
#      max_time = 300
#      dot_matrix = np.empty((max_time + 1, 224, 2))
#      for i in range(290, max_time + 1):
#          print(f'{i}s')
#          read_from_dis_excel(i)
#          print(all_cross_check(dot_matrix[i]))

def cal_nearest(space):
    #print(t_to_dis(0, space))
    delta = 10
    crash = 280
    epsilon = 0.01
    while delta >= epsilon:
        low = 230
        high = crash
        time = low
        while time < high:
            time = time + delta
            matrix = t_to_dis(time, space)
            print(time)
            #print(all_cross_check(matrix))
            if all_cross_check(matrix):
                crash = time if time < crash else crash
                break
        delta /= 10
    print(crash)

#cal_nearest(45)

def local_check():
    for t in range(40950, 41200, 5):
        time = t / 100
        matrix = t_to_dis(time, 55)
        #print(t_to_dis(time))
        print(time)
        print(all_cross_check(matrix))

max = 30
w = 0.5
c1 = 0.5
c2 = 1
max_time = 500
def pso_cal_min_distance(space, n, c1, c2):
    partical = []
    partical_v = []
    pbest    = []
    gbest    = (max_time, True)
    for i in range(n):
        print(f'initial:{i}/{n}')
        partical.append(random.random() * max_time)
        partical_v.append(random.random())
        pbest.append((partical[i], all_cross_check(t_to_dis(partical[i], space))))
        if pbest[i][0] < gbest[0] and pbest[i][1]:
            gbest = pbest[i]

    for j in range(max):
        for i in range(n):
            partical_v[i] = w * partical_v[i] + c1 * random.random() * (
                        pbest[i][0] - partical[i]) + c2 * random.random() * (gbest[0] - partical[i])
            partical[i] = partical[i] + partical_v[i]

            if partical[i] > 500:
                partical[i] = 500
            if partical[i] < 0:
                partical[i] = 0

            cross = all_cross_check(t_to_dis(partical[i], space))
            if not pbest[i][1] and cross:
                pbest[i] = (partical[i], True)

            elif pbest[i][1] and cross:
                pbest[i] = pbest[i] if pbest[i][0] < partical[i] else (partical[i], cross)

            if pbest[i][0] < gbest[0] and pbest[i][1]:
                gbest = pbest[i]

            # if pbest[i][0] > gbest[0]:
            #     partical[i] = gbest[0] * random.random()
            #     pbest[i] = partical[i], all_cross_check(t_to_dis(partical[i], space)) # if not best, go to a random smaller position

            print(f'{i}, {pbest[i]}, {gbest}')
            print(f'{space}, {t_to_theta(gbest[0], D=space / 100) / 2 / np.pi * space}')
            # if t_to_theta(gbest[0], D=space / 100) / 2 / np.pi * space > 450:
            #     return False # cannot enter

    return True # can enter

smallest = 46
biggist  = 50

# for i in range(smallest, biggist + 1, 1):
#     pso_cal_min_distance(i, 10, c1, c2)
#     pso_cal_min_distance(i + 0.5, 10, c1, c2)
#
# smallest = 545
# biggist  = 555
#
# for i in range(smallest, biggist + 1, 1):
#     pso_cal_min_distance(i / 10, 10, c1, c2)

# pso_cal_min_distance(55, 10, c1, c2)
#
# smallest = 45
# biggist  = 55
# space = (biggist + smallest) / 2
# for i in range(10):
#     print(space)
#     if pso_cal_min_distance(space, 10, c1, c2):
#         biggist = space
#     else :
#         smallest = space
#     space = (biggist + smallest) / 2
